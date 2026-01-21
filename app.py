from io import BytesIO
from flask import Flask, render_template, request, send_file, jsonify, session, redirect, url_for
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import secrets
import sqlite3

app = Flask(__name__)
# Vercel 배포를 위한 시크릿 키 (환경 변수 또는 랜덤 생성)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(16))

# 전역 변수: 생성된 엑셀 워크북 저장 (세션별)
# 세션 ID를 키로 사용하여 각 사용자별로 독립적인 워크북 관리
workbooks = {}

#############################################
# QUAZ 갤러리 (간단 게시판) - SQLite 기반
#############################################

def _quaz_db_path():
    instance_dir = os.path.join(app.root_path, "instance")
    os.makedirs(instance_dir, exist_ok=True)
    return os.path.join(instance_dir, "quaz_gallery.db")


def quaz_get_db():
    conn = sqlite3.connect(_quaz_db_path())
    conn.row_factory = sqlite3.Row
    return conn


def quaz_init_db():
    conn = quaz_get_db()
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS posts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                content TEXT NOT NULL,
                media_url TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                views INTEGER NOT NULL DEFAULT 0
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS comments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                post_id INTEGER NOT NULL,
                parent_id INTEGER,
                author TEXT,
                content TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(post_id) REFERENCES posts(id)
            )
            """
        )
        conn.commit()
    finally:
        conn.close()


def _now_iso():
    return datetime.now().isoformat(timespec="seconds")


@app.before_request
def _ensure_quaz_db():
    quaz_init_db()


@app.route("/", methods=["GET"])
def quaz_index():
    """QUAZ 갤러리 - 게시물 목록 (최신글 내림차순)"""
    q = (request.args.get("q") or "").strip()
    conn = quaz_get_db()
    try:
        if q:
            like = f"%{q}%"
            posts = conn.execute(
                """
                SELECT id, title, created_at, views
                FROM posts
                WHERE title LIKE ? OR content LIKE ?
                ORDER BY datetime(created_at) DESC, id DESC
                """,
                (like, like),
            ).fetchall()
        else:
            posts = conn.execute(
                """
                SELECT id, title, created_at, views
                FROM posts
                ORDER BY datetime(created_at) DESC, id DESC
                """
            ).fetchall()
    finally:
        conn.close()
    return render_template("quaz_index.html", posts=posts, q=q)


@app.route("/write", methods=["GET", "POST"])
def quaz_write():
    """글쓰기 페이지 (이미지/동영상 링크 첨부 가능)"""
    if request.method == "POST":
        title = (request.form.get("title") or "").strip()
        content = (request.form.get("content") or "").strip()
        media_url = (request.form.get("media_url") or "").strip() or None

        if not title or not content:
            return render_template(
                "quaz_form.html",
                mode="create",
                error="제목과 내용을 입력해주세요.",
                post={"title": title, "content": content, "media_url": media_url or ""},
            )

        conn = quaz_get_db()
        try:
            conn.execute(
                "INSERT INTO posts(title, content, media_url, created_at) VALUES(?,?,?,?)",
                (title, content, media_url, _now_iso()),
            )
            conn.commit()
        finally:
            conn.close()

        return redirect(url_for("quaz_index"))

    return render_template("quaz_form.html", mode="create", post={"title": "", "content": "", "media_url": ""})


@app.route("/post/<int:post_id>", methods=["GET"])
def quaz_post_detail(post_id: int):
    """게시물 상세 + 조회수 증가 + 댓글/대댓글"""
    conn = quaz_get_db()
    try:
        conn.execute("UPDATE posts SET views = views + 1 WHERE id = ?", (post_id,))
        conn.commit()

        post = conn.execute("SELECT * FROM posts WHERE id = ?", (post_id,)).fetchone()
        if not post:
            return "게시물을 찾을 수 없습니다.", 404

        comments = conn.execute(
            """
            SELECT id, post_id, parent_id, author, content, created_at
            FROM comments
            WHERE post_id = ?
            ORDER BY datetime(created_at) ASC, id ASC
            """,
            (post_id,),
        ).fetchall()
    finally:
        conn.close()

    by_parent = {}
    for c in comments:
        by_parent.setdefault(c["parent_id"], []).append(c)

    def build_tree(parent_id):
        items = []
        for c in by_parent.get(parent_id, []):
            items.append({"comment": c, "replies": build_tree(c["id"])})
        return items

    comment_tree = build_tree(None)
    return render_template("quaz_post_detail.html", post=post, comment_tree=comment_tree)


@app.route("/post/<int:post_id>/comment", methods=["POST"])
def quaz_add_comment(post_id: int):
    author = (request.form.get("author") or "").strip() or None
    content = (request.form.get("content") or "").strip()
    parent_id_raw = (request.form.get("parent_id") or "").strip()
    parent_id = int(parent_id_raw) if parent_id_raw.isdigit() else None

    if not content:
        return redirect(url_for("quaz_post_detail", post_id=post_id))

    conn = quaz_get_db()
    try:
        conn.execute(
            "INSERT INTO comments(post_id, parent_id, author, content, created_at) VALUES(?,?,?,?,?)",
            (post_id, parent_id, author, content, _now_iso()),
        )
        conn.commit()
    finally:
        conn.close()

    return redirect(url_for("quaz_post_detail", post_id=post_id))


@app.route("/post/<int:post_id>/edit", methods=["GET", "POST"])
def quaz_edit_post(post_id: int):
    conn = quaz_get_db()
    try:
        post = conn.execute("SELECT * FROM posts WHERE id = ?", (post_id,)).fetchone()
        if not post:
            return "게시물을 찾을 수 없습니다.", 404

        if request.method == "POST":
            title = (request.form.get("title") or "").strip()
            content = (request.form.get("content") or "").strip()
            media_url = (request.form.get("media_url") or "").strip() or None

            if not title or not content:
                return render_template(
                    "quaz_form.html",
                    mode="edit",
                    error="제목과 내용을 입력해주세요.",
                    post={"id": post_id, "title": title, "content": content, "media_url": media_url or ""},
                )

            conn.execute(
                "UPDATE posts SET title=?, content=?, media_url=?, updated_at=? WHERE id=?",
                (title, content, media_url, _now_iso(), post_id),
            )
            conn.commit()
            return redirect(url_for("quaz_post_detail", post_id=post_id))
    finally:
        conn.close()

    return render_template(
        "quaz_form.html",
        mode="edit",
        post={"id": post["id"], "title": post["title"], "content": post["content"], "media_url": post["media_url"] or ""},
    )


@app.route("/post/<int:post_id>/delete", methods=["POST"])
def quaz_delete_post(post_id: int):
    conn = quaz_get_db()
    try:
        conn.execute("DELETE FROM comments WHERE post_id = ?", (post_id,))
        conn.execute("DELETE FROM posts WHERE id = ?", (post_id,))
        conn.commit()
    finally:
        conn.close()
    return redirect(url_for("quaz_index"))


@app.route("/trends", methods=["GET"])
def quaz_trends():
    """요즘 트렌드 - 조회수 높은 게시물 목록"""
    conn = quaz_get_db()
    try:
        posts = conn.execute(
            """
            SELECT id, title, created_at, views
            FROM posts
            ORDER BY views DESC, datetime(created_at) DESC
            LIMIT 50
            """
        ).fetchall()
    finally:
        conn.close()
    return render_template("quaz_trends.html", posts=posts)


@app.route("/admin", methods=["GET"])
def quaz_admin():
    return render_template("quaz_admin.html")


def get_or_create_workbook(session_id):
    """세션별 워크북 가져오기 또는 생성"""
    if session_id not in workbooks:
        # 새 워크북 생성
        wb = openpyxl.Workbook()
        # 기본 시트 삭제
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        workbooks[session_id] = wb
    return workbooks[session_id]


def read_data_from_excel(number_input):
    """
    CON-A DB1.xlsx에서 지정된 번호(1 또는 2)에 해당하는 데이터를 읽어옴
    
    A의 정보(번호 1):
    - 대가 시트: 대가 시트의 3, 4, 5번 줄
    - 집계 시트: 대가 시트의 4번 줄
    
    B의 정보(번호 2):
    - 대가 시트: 참조 시트의 7, 8, 9번 줄
    - 집계 시트: 대가 시트의 5번 줄
    """
    excel_path = "data/CON-A DB1.xlsx"
    
    # openpyxl로 원본 파일 읽기
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    daega_data = []  # 대가 시트 데이터
    jipgye_data = []  # 집계 시트 데이터
    
    if number_input == '1':  # A의 정보
        # 대가 시트 찾기
        daega_sheet_name = None
        for name in wb.sheetnames:
            if '대가' in name:
                daega_sheet_name = name
                break
        if daega_sheet_name is None:
            daega_sheet_name = wb.sheetnames[0]  # 첫 번째 시트 사용
        
        # 대가 시트: 대가 시트의 3, 4, 5번 줄
        if daega_sheet_name:
            ws_daega = wb[daega_sheet_name]
            for row_num in [3, 4, 5]:
                row_data = {}
                has_data = False
                for col_idx in range(1, 12):  # A부터 K까지 (11개 컬럼)
                    col_letter = get_column_letter(col_idx)
                    cell = ws_daega.cell(row=row_num, column=col_idx)
                    value = cell.value if cell.value is not None else ''
                    row_data[col_letter] = value
                    if value != '':
                        has_data = True
                if has_data:  # 데이터가 있는 행만 추가
                    row_data['원본시트'] = daega_sheet_name
                    row_data['원본행'] = row_num
                    daega_data.append(row_data)
        
        # 집계 시트: 첫 번째 시트에서 A열=1, B열='A'인 행 찾기 (A의 정보)
        # 사용자 요구사항: A,5,2.5,12.5,3,15,4,20.9.5,47.5
        # 실제 Excel 데이터: 첫 번째 시트 Row 4 = [1, 'A', 5, 2.5, 12.5, 3, 15, 4, 20, 9.5, 47.5, None]
        jipgye_sheet_name = wb.sheetnames[0]  # 첫 번째 시트 사용
        ws_jipgye = wb[jipgye_sheet_name]
        # 첫 번째 시트에서 A열=1, B열='A'인 행 찾기
        target_row = None
        for row_num in range(1, 15):
            cell_a = ws_jipgye.cell(row=row_num, column=1)
            cell_b = ws_jipgye.cell(row=row_num, column=2)
            if cell_a.value == 1 and cell_b and str(cell_b.value).strip().upper() == 'A':
                target_row = row_num
                break
        
        # 찾지 못했으면 4번 줄 사용
        if target_row is None:
            target_row = 4
        
        row_data = {}
        has_data = False
        # A부터 L까지 (12개 컬럼) - 실제로는 A부터 K까지 데이터가 있음
        for col_idx in range(1, 13):
            col_letter = get_column_letter(col_idx)
            cell = ws_jipgye.cell(row=target_row, column=col_idx)
            value = cell.value if cell.value is not None else ''
            row_data[col_letter] = value
            if value != '':
                has_data = True
        if has_data:  # 데이터가 있는 경우만 추가
            row_data['원본시트'] = jipgye_sheet_name
            row_data['원본행'] = target_row
            jipgye_data.append(row_data)
    
    elif number_input == '2':  # B의 정보
        # 대가 시트: 참조 시트의 7, 8, 9번 줄
        # 시트 이름 확인
        sheet_name = None
        for name in wb.sheetnames:
            if '참조' in name:
                sheet_name = name
                break
        
        if sheet_name is None and len(wb.sheetnames) > 1:
            # 두 번째 시트가 참조 시트일 가능성
            sheet_name = wb.sheetnames[1]
        
        if sheet_name:
            ws_daega = wb[sheet_name]
            for row_num in [7, 8, 9]:
                row_data = {}
                has_data = False
                for col_idx in range(1, 12):  # A부터 K까지
                    col_letter = get_column_letter(col_idx)
                    cell = ws_daega.cell(row=row_num, column=col_idx)
                    value = cell.value if cell.value is not None else ''
                    row_data[col_letter] = value
                    if value != '':
                        has_data = True
                if has_data:  # 데이터가 있는 행만 추가
                    row_data['원본시트'] = sheet_name
                    row_data['원본행'] = row_num
                    daega_data.append(row_data)
        
        # 집계 시트: 첫 번째 시트에서 A열=2, B열='B'인 행 찾기 (B의 정보)
        # 사용자 요구사항: B,10,21,210,4.5,45,11,110,36.5,365
        # 실제 Excel 데이터: 첫 번째 시트 Row 5 = [2, 'B', 10, 21, 210, 4.5, 45, 11, 110, 36.5, 365, None]
        jipgye_sheet_name = wb.sheetnames[0]  # 첫 번째 시트 사용
        ws_jipgye = wb[jipgye_sheet_name]
        # 첫 번째 시트에서 A열=2, B열='B'인 행 찾기
        target_row = None
        for row_num in range(1, 15):
            cell_a = ws_jipgye.cell(row=row_num, column=1)
            cell_b = ws_jipgye.cell(row=row_num, column=2)
            if cell_a.value == 2 and cell_b and str(cell_b.value).strip().upper() == 'B':
                target_row = row_num
                break
        
        # 찾지 못했으면 5번 줄 사용
        if target_row is None:
            target_row = 5
        
        row_data = {}
        has_data = False
        # A부터 L까지 (12개 컬럼) - 실제로는 A부터 K까지 데이터가 있음
        for col_idx in range(1, 13):
            col_letter = get_column_letter(col_idx)
            cell = ws_jipgye.cell(row=target_row, column=col_idx)
            value = cell.value if cell.value is not None else ''
            row_data[col_letter] = value
            if value != '':
                has_data = True
        if has_data:  # 데이터가 있는 경우만 추가
            row_data['원본시트'] = jipgye_sheet_name
            row_data['원본행'] = target_row
            jipgye_data.append(row_data)
    
    return daega_data, jipgye_data


def create_daega_sheet(wb, data_list, sheet_name="대가"):
    """대가 시트 생성 또는 기존 시트에 데이터 추가"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 기존 데이터의 마지막 행 찾기 (헤더가 있으면 그 다음부터)
        if ws.max_row > 0:
            start_row = ws.max_row + 1
        else:
            start_row = 2  # 헤더 행 다음
            # 헤더 작성
            headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
    else:
        ws = wb.create_sheet(sheet_name)
        start_row = 2
        # 헤더 작성 - A부터 K까지
        headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    
    # 데이터 추가
    for row_data in data_list:
        # A부터 K까지 최대 11개 컬럼 처리
        for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'], start=1):
            value = row_data.get(col_letter, '')
            # 빈 문자열이 아닌 경우에만 값 설정
            if value != '':
                ws.cell(row=start_row, column=col_idx, value=value)
        start_row += 1
    
    return ws


def create_jipgye_sheet(wb, data_list, sheet_name="집계"):
    """집계 시트 생성 또는 기존 시트에 데이터 추가"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 기존 데이터의 마지막 행 찾기 (헤더가 있으면 그 다음부터)
        if ws.max_row > 0:
            start_row = ws.max_row + 1
        else:
            start_row = 2  # 헤더 행 다음
            # 헤더 작성 - B부터 시작 (A열 제외)
            headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
    else:
        ws = wb.create_sheet(sheet_name)
        start_row = 2
        # 헤더 작성 - A부터 L까지 (최대 12개 컬럼)
        headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    
    # 집계 데이터 추가 - A열(숫자)을 제외하고 B열부터 시작
    for row_data in data_list:
        # B열부터 L까지 저장 (A열 제외)
        # 원본 데이터: A열=1 또는 2, B열='A' 또는 'B', C열부터 데이터
        # 집계 시트: B열부터 시작하여 저장
        col_letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        for col_idx, col_letter in enumerate(col_letters, start=2):  # B열부터 시작 (컬럼 인덱스 2)
            value = row_data.get(col_letter, '')
            # 빈 문자열이 아니고 None이 아닌 경우에만 값 설정
            if value != '' and value is not None:
                ws.cell(row=start_row, column=col_idx, value=value)
        start_row += 1
    
    return ws


@app.route("/excel", methods=["GET", "POST"])
def index():
    """(기존) CON-A 메인 페이지"""
    session_id = session.get('session_id')
    if not session_id:
        session_id = datetime.now().strftime('%Y%m%d%H%M%S%f')
        session['session_id'] = session_id
    
    message = None
    error = None
    
    if request.method == "POST":
        number_input = request.form.get("number_input", "").strip()
        
        if number_input in ['1', '2']:
            # 엑셀에서 데이터 읽기
            daega_data, jipgye_data = read_data_from_excel(number_input)
            
            # 워크북 가져오기 또는 생성
            wb = get_or_create_workbook(session_id)
            
            # 대가 시트 생성 또는 추가 (항상 생성)
            create_daega_sheet(wb, daega_data, "대가")
            
            # 집계 시트 생성 또는 추가 (항상 생성)
            create_jipgye_sheet(wb, jipgye_data, "집계")
            
            total_rows = len(daega_data) + len(jipgye_data)
            if total_rows > 0:
                message = f"번호 {number_input} 데이터가 추가되었습니다. 대가 시트: {len(daega_data)}개 행, 집계 시트: {len(jipgye_data)}개 행이 추가되었습니다."
            else:
                error = f"번호 {number_input}에 해당하는 데이터를 찾을 수 없습니다."
        else:
            error = "번호는 1 또는 2만 입력 가능합니다."
    
    # 현재 워크북의 시트 정보 및 데이터 가져오기
    wb = get_or_create_workbook(session_id)
    sheet_info = {}
    sheet_data = {}  # 각 시트의 실제 데이터
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_count = ws.max_row - 1 if ws.max_row > 1 else 0  # 헤더 제외
        sheet_info[sheet_name] = {'row_count': row_count}
        
        # 시트 데이터 읽기
        sheet_rows = []
        headers = []
        
        if ws.max_row > 1:
            # 헤더 읽기
            max_col = ws.max_column if ws.max_column > 0 else 12  # 최소 12개 컬럼
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col_idx)
                value = cell.value if cell.value is not None else ''
                headers.append(value)
            
            # 데이터 행 읽기
            for row_idx in range(2, ws.max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value if cell.value is not None else ''
                    row_data.append(value)
                sheet_rows.append(row_data)
        
        sheet_data[sheet_name] = {
            'headers': headers,
            'rows': sheet_rows
        }
    
    return render_template(
        "index.html",
        message=message,
        error=error,
        sheet_info=sheet_info,
        sheet_data=sheet_data,
        session_id=session_id
    )


@app.route("/download")
def download():
    """엑셀 파일 다운로드"""
    session_id = session.get('session_id')
    
    if not session_id or session_id not in workbooks:
        return "생성된 데이터가 없습니다. 먼저 번호를 입력해주세요.", 400
    
    wb = workbooks[session_id]
    
    if len(wb.sheetnames) == 0:
        return "생성된 시트가 없습니다. 먼저 번호를 입력해주세요.", 400
    
    # 메모리 버퍼에 엑셀 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"CON-A_결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/clear", methods=["POST"])
def clear():
    """현재 세션의 워크북 초기화"""
    session_id = session.get('session_id')
    
    if session_id and session_id in workbooks:
        del workbooks[session_id]
    
    return redirect(url_for('index'))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)

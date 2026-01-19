from io import BytesIO
from typing import List

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, send_file

app = Flask(__name__)


def load_excel_sheets(path: str) -> List[str]:
    """
    CON-A DB1.xlsx에 존재하는 시트 이름 목록을 반환한다.
    """
    xls = pd.ExcelFile(path)
    return xls.sheet_names


def query_data_multi_sheet(selected_char: str):
    """
    선택문자를 기준으로 집계와 대가 시트에서 데이터를 조회한다.
    원본 파일을 그대로 불러와서 표시한다.
    """
    excel_path = "data/CON-A DB1.xlsx"
    
    # 모든 시트에서 데이터 조회
    all_rows = []
    sheet_names = load_excel_sheets(excel_path)
    
    # 계산된 값을 가져오기 위해 data_only=True로 읽기
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)
    
    for sheet_name in sheet_names:
        try:
            # 헤더 행이 3번째 행(인덱스 2)에 있으므로 header=2로 설정
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=2)
            df = df.dropna(how='all')
            
            # 선택문자 필터링 (B열, 인덱스 1 = 두 번째 컬럼)
            # 대소문자 구분 없이 비교
            if selected_char and len(df.columns) > 1:
                char_col = df.columns[1]  # B열
                selected_char_upper = str(selected_char).strip().upper()
                df = df[df[char_col].astype(str).str.strip().str.upper() == selected_char_upper]
            
            ws_data = wb_data[sheet_name]
            
            # 각 행을 처리 - 원본 구조 그대로
            for idx, (_, row) in enumerate(df.iterrows()):
                new_row = {}
                new_row['시트'] = sheet_name
                
                # 원본 파일의 실제 행 번호 (헤더가 3번째 행이므로 +3)
                actual_row_idx = idx + 4  # pandas 인덱스 + 헤더 3행 + 1
                
                # 모든 컬럼을 원본 그대로 저장 (A, B, C, D, E, F, G, H, I, J, K...)
                for col_idx in range(len(df.columns)):
                    col_letter = get_column_letter(col_idx + 1)
                    cell_value = row.iloc[col_idx] if col_idx < len(row) else ''
                    new_row[col_letter] = cell_value
                
                # 행번호 (A열)
                if len(df.columns) > 0:
                    new_row['행번호'] = row.iloc[0] if pd.notna(row.iloc[0]) else ''
                
                # 선택문자 (B열)
                if len(df.columns) > 1:
                    new_row['선택문자'] = row.iloc[1] if pd.notna(row.iloc[1]) else ''
                
                # 숫자 (C열) - 입력 가능하게 하기 위해 별도 저장
                if len(df.columns) > 2:
                    new_row['숫자'] = row.iloc[2] if pd.notna(row.iloc[2]) else ''
                
                # 원본 값 저장 (D, F, H열의 계산된 값)
                if len(df.columns) > 3:
                    d_cell = ws_data.cell(row=actual_row_idx, column=4)
                    try:
                        new_row['_가_원본'] = float(d_cell.value) if d_cell.value else 0
                    except:
                        new_row['_가_원본'] = 0
                
                if len(df.columns) > 5:
                    f_cell = ws_data.cell(row=actual_row_idx, column=6)
                    try:
                        new_row['_나_원본'] = float(f_cell.value) if f_cell.value else 0
                    except:
                        new_row['_나_원본'] = 0
                
                if len(df.columns) > 7:
                    h_cell = ws_data.cell(row=actual_row_idx, column=8)
                    try:
                        new_row['_다_원본'] = float(h_cell.value) if h_cell.value else 0
                    except:
                        new_row['_다_원본'] = 0
                
                all_rows.append(new_row)
        except Exception as e:
            print(f"시트 {sheet_name} 로드 오류: {e}")
            continue
    
    if not all_rows:
        return pd.DataFrame()
    
    result_df = pd.DataFrame(all_rows)
    return result_df


@app.route("/", methods=["GET", "POST"])
def index():
    excel_path = "data/CON-A DB1.xlsx"

    # 시트 목록 로드
    try:
        sheet_names = load_excel_sheets(excel_path)
    except Exception as e:
        # 엑셀 로드 실패 시 빈 상태로 렌더링
        return render_template(
            "index.html",
            error=f"엑셀 파일을 로드할 수 없습니다: {e}",
            sheet_names=[],
            current_sheet="",
            selected_char="",
            rows=[],
            columns=[],
            column_labels=[],
            sheet_groups={},
        )

    # 선택문자 입력 시 모든 시트에서 조회
    selected_input = request.form.get("selected_char", "").strip()
    
    # 숫자를 선택문자로 매핑: 1 → A, 2 → B
    char_mapping = {
        '1': 'A',
        '2': 'B',
    }
    
    # 입력값이 숫자면 선택문자로 변환, 아니면 그대로 사용
    if selected_input in char_mapping:
        selected_char = char_mapping[selected_input]
    else:
        selected_char = selected_input
    
    # 선택문자가 있으면 모든 시트에서 조회, 없으면 첫 번째 시트만 표시
    if selected_char:
        # 선택문자가 있으면 모든 시트에서 조회하고 컬럼 통합
        df = query_data_multi_sheet(selected_char)
        current_sheet = "전체"
    else:
        # 선택문자가 없으면 첫 번째 시트만 표시하되 컬럼 통합 적용
        current_sheet = request.form.get("sheet_name") or sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=current_sheet, header=2)
        df = df.dropna(how='all')
        
        # 컬럼 통합: E, G, I, K 열 삭제하고 D+E를 가열, F+G를 나열, H+I를 다열로 통합
        result_rows = []
        for _, row in df.iterrows():
            new_row = {}
            new_row['시트'] = current_sheet
            new_row['행번호'] = row.iloc[0] if len(df.columns) > 0 and pd.notna(row.iloc[0]) else ''
            new_row['선택문자'] = row.iloc[1] if len(df.columns) > 1 and pd.notna(row.iloc[1]) else ''
            new_row['숫자'] = row.iloc[2] if len(df.columns) > 2 and pd.notna(row.iloc[2]) else ''
            
            # 가열: D+E
            if len(df.columns) > 4:
                d_val = row.iloc[3] if pd.notna(row.iloc[3]) else 0
                e_val = row.iloc[4] if pd.notna(row.iloc[4]) else 0
                try:
                    d_num = float(d_val) if d_val != '' else 0
                    new_row['가열'] = d_num + float(e_val) if e_val != '' else d_num
                    new_row['_가_원본'] = d_num
                except:
                    new_row['가열'] = f"{d_val} + {e_val}"
                    new_row['_가_원본'] = 0
            
            # 나열: F+G
            if len(df.columns) > 6:
                f_val = row.iloc[5] if pd.notna(row.iloc[5]) else 0
                g_val = row.iloc[6] if pd.notna(row.iloc[6]) else 0
                try:
                    f_num = float(f_val) if f_val != '' else 0
                    new_row['나열'] = f_num + float(g_val) if g_val != '' else f_num
                    new_row['_나_원본'] = f_num
                except:
                    new_row['나열'] = f"{f_val} + {g_val}"
                    new_row['_나_원본'] = 0
            
            # 다열: H+I
            if len(df.columns) > 8:
                h_val = row.iloc[7] if pd.notna(row.iloc[7]) else 0
                i_val = row.iloc[8] if pd.notna(row.iloc[8]) else 0
                try:
                    h_num = float(h_val) if h_val != '' else 0
                    new_row['다열'] = h_num + float(i_val) if i_val != '' else h_num
                    new_row['_다_원본'] = h_num
                except:
                    new_row['다열'] = f"{h_val} + {i_val}"
                    new_row['_다_원본'] = 0
            
            # 합계: J
            if len(df.columns) > 9:
                new_row['합계'] = row.iloc[9] if pd.notna(row.iloc[9]) else ''
            
            result_rows.append(new_row)
        
        df = pd.DataFrame(result_rows)

    # 템플릿으로 넘길 데이터 준비
    rows = df.to_dict(orient="records")
    columns = list(df.columns)
    
    # 시트별로 그룹화 (선택문자가 있고 시트 컬럼이 있을 때)
    sheet_groups = {}
    if selected_char and '시트' in columns and rows:
        for row in rows:
            sheet_name = row.get('시트', '')
            if sheet_name:
                if sheet_name not in sheet_groups:
                    sheet_groups[sheet_name] = []
                sheet_groups[sheet_name].append(row)
    
    # 컬럼명을 원본 그대로 표시 (A, B, C, D, E, F, G, H, I, J, K...)
    column_labels = []
    for col in columns:
        if col == '시트' or col.startswith('_'):
            continue  # 시트와 내부 컬럼은 제외
        elif col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
            column_labels.append(col)
        elif col == '행번호':
            column_labels.append('A')
        elif col == '선택문자':
            column_labels.append('B')
        elif col == '숫자':
            column_labels.append('C')
        else:
            column_labels.append(str(col))

    return render_template(
        "index.html",
        sheet_names=sheet_names,
        current_sheet=current_sheet,
        selected_char=selected_char,
        rows=rows,
        columns=columns,
        column_labels=column_labels,
        sheet_groups=sheet_groups,
        error=None,
    )


@app.route("/download", methods=["POST"])
def download():
    """
    원본 data/CON-A DB1.xlsx 파일을 읽어서 수량만 수정하고
    계산된 값을 원본 위치에 반영하여 다운로드한다.
    """
    selected_input = request.form.get("selected_char", "").strip()

    if not selected_input:
        return "선택문자를 입력해주세요.", 400

    # 숫자를 선택문자로 매핑: 1 → A, 2 → B
    char_mapping = {
        '1': 'A',
        '2': 'B',
    }
    
    # 입력값이 숫자면 선택문자로 변환, 아니면 그대로 사용
    if selected_input in char_mapping:
        selected_char = char_mapping[selected_input]
    else:
        selected_char = selected_input

    # 웹에서 입력한 숫자 값들을 받아옴 (JSON 형식)
    import json
    number_values_json = request.form.get("number_values", "{}")
    try:
        number_values = json.loads(number_values_json)
    except:
        number_values = {}

    excel_path = "data/CON-A DB1.xlsx"
    
    # 원본 엑셀 파일을 그대로 복사 (아무것도 건드리지 않음)
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    # 계산된 값을 가져오기 위해 data_only=True로도 읽기
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)
    
    # 각 시트 처리
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]  # 계산된 값을 가져오기 위한 시트
        
        # 헤더 행이 3번째 행이므로 데이터는 4번째 행부터
        # 선택문자로 필터링된 행만 처리
        matched_row_index = 0  # 웹에서 표시된 행 인덱스 (0부터 시작)
        for row_idx in range(4, ws.max_row + 1):
            # B열(컬럼 2)이 선택문자
            char_cell = ws.cell(row=row_idx, column=2)
            if char_cell.value:
                char_value = str(char_cell.value).strip().upper()
                if char_value == selected_char.upper():
                    # 웹에서 입력한 숫자 값 가져오기
                    # 키 형식: "시트명_행인덱스" (0부터 시작)
                    sheet_key = f"{sheet_name}_{matched_row_index}"
                    if sheet_key in number_values:
                        number_value = float(number_values[sheet_key])
                        
                        # C열(컬럼 3)이 숫자 칸 - 숫자 칸만 수정
                        number_cell = ws.cell(row=row_idx, column=3)
                        number_cell.value = number_value
                        
                        # D열(컬럼 4)이 가 원본 - 계산된 값 가져오기
                        d_cell_data = ws_data.cell(row=row_idx, column=4)
                        try:
                            d_value = float(d_cell_data.value) if d_cell_data.value else 0
                        except:
                            d_value = 0
                        
                        # E열(컬럼 5)에 가열 결과: 숫자 × D열 (원본 위치에 출력)
                        e_cell = ws.cell(row=row_idx, column=5)
                        e_cell.value = number_value * d_value
                        
                        # F열(컬럼 6)이 나 원본 - 계산된 값 가져오기
                        f_cell_data = ws_data.cell(row=row_idx, column=6)
                        try:
                            f_value = float(f_cell_data.value) if f_cell_data.value else 0
                        except:
                            f_value = 0
                        
                        # G열(컬럼 7)에 나열 결과: 숫자 × F열 (원본 위치에 출력)
                        g_cell = ws.cell(row=row_idx, column=7)
                        g_cell.value = number_value * f_value
                        
                        # H열(컬럼 8)이 다 원본 - 계산된 값 가져오기
                        h_cell_data = ws_data.cell(row=row_idx, column=8)
                        try:
                            h_value = float(h_cell_data.value) if h_cell_data.value else 0
                        except:
                            h_value = 0
                        
                        # I열(컬럼 9)에 다열 결과: 숫자 × H열 (원본 위치에 출력)
                        i_cell = ws.cell(row=row_idx, column=9)
                        i_cell.value = number_value * h_value
                    
                    matched_row_index += 1

    # 메모리 버퍼에 엑셀 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "CON-A_result.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    # 개발 편의를 위해 debug 모드 사용 (배포 시 False로 변경)
    try:
        app.run(host="0.0.0.0", port=5000, debug=True)
    except OSError as e:
        if "Address already in use" in str(e):
            print("\n포트 5000이 이미 사용 중입니다.")
            print("다른 포트로 실행하려면 코드의 port 값을 변경하세요.")
            print("또는 포트 5000을 사용하는 프로그램을 종료하세요.\n")
        raise